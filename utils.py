import openpyxl
import csv
from datetime import datetime
from models import Transaction, Loan, Distribution

# 其他函数保持不变...

def import_from_csv(file_path):
    """从CSV文件导入数据"""
    transactions = []
    
    with open(file_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # 解析日期
            date_str = row.get('日期', '').strip()
            if not date_str:
                continue
            
            try:
                # 处理 2025/1/9 格式
                date = datetime.strptime(date_str, '%Y/%m/%d').date()
            except ValueError:
                try:
                    # 处理 2025-1-9 格式
                    date = datetime.strptime(date_str, '%Y-%m-%d').date()
                except ValueError:
                    continue
            
            # 解析收支分类
            type_str = row.get('收支分类', '').strip()
            if type_str == '入账':
                trans_type = '入账'
            elif type_str == '出账':
                trans_type = '出账'
            else:
                continue
            
            # 解析明细分类
            category = row.get('明细分类', '').strip()
            
            # 解析摘要
            description = row.get('摘要', '').strip()
            
            # 解析金额
            amount_str = row.get('金额', '').strip()
            if not amount_str:
                continue
            
            try:
                # 移除逗号和引号
                amount_str = amount_str.replace(',', '').replace('"', '')
                amount = float(amount_str)
                # 确保金额为正数
                amount = abs(amount)
            except ValueError:
                continue
            
            # 解析备注
            remark = row.get(' 备注 ', '').strip()
            
            # 创建交易记录
            transaction = Transaction(
                date=date,
                type=trans_type,
                category=category,
                amount=amount,
                description=description,
                remark=remark,
                month=date.month,
                year=date.year
            )
            transactions.append(transaction)
    
    return {
        'transactions': transactions,
        'loans': [],
        'distributions': []
    }

def safe_float(value):
    """安全地将值转换为浮点数"""
    if value is None:
        return 0
    try:
        return float(value)
    except (ValueError, TypeError):
        return 0

def parse_date(date_value):
    if date_value is None:
        return None
    
    if isinstance(date_value, datetime):
        return date_value.date()
    
    if isinstance(date_value, str):
        try:
            return datetime.strptime(date_value, '%Y-%m-%d').date()
        except ValueError:
            try:
                return datetime.strptime(date_value, '%Y.%m.%d').date()
            except ValueError:
                try:
                    month, day = map(float, date_value.split('.'))
                    year = 2025  # 默认年份
                    return datetime(year, int(month), int(day)).date()
                except:
                    return None
    
    if isinstance(date_value, (int, float)):
        try:
            return datetime.fromordinal(datetime(1900, 1, 1).toordinal() + int(date_value) - 2).date()
        except:
            return None
    
    return None

def import_from_excel(file_path):
    transactions = []
    loans = []
    distributions = []
    
    wb = openpyxl.load_workbook(file_path)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        if sheet_name in ['总账', '小张', '王永雄']:
            for row in ws.iter_rows(min_row=4):
                if len(row) > 0 and row[0].value:
                    date = parse_date(row[0].value)
                    if date:
                        # 修正：第2列是摘要，第3列是金额
                        description = row[1].value if len(row) > 1 and row[1].value else ''
                        amount = safe_float(row[2].value) if len(row) > 2 else 0
                        category = description  # 使用摘要作为分类
                        
                        if amount > 0:
                            transaction = Transaction(
                                date=date,
                                type='支出',
                                category=str(category),
                                amount=amount,
                                description=str(description),
                                month=date.month,
                                year=date.year
                            )
                            transactions.append(transaction)
                
                if len(row) > 4 and row[4].value:
                    date = parse_date(row[4].value)
                    if date:
                        # 修正：第6列是摘要，第7列是金额
                        description = row[5].value if len(row) > 5 and row[5].value else ''
                        amount = safe_float(row[6].value) if len(row) > 6 else 0
                        category = description  # 使用摘要作为分类
                        
                        if amount > 0:
                            transaction = Transaction(
                                date=date,
                                type='入账',
                                category=str(category),
                                amount=amount,
                                description=str(description),
                                month=date.month,
                                year=date.year
                            )
                            transactions.append(transaction)
                
                if len(row) > 8 and row[8].value:
                    date = parse_date(row[8].value)
                    if date:
                        # 修正：第10列是摘要，第11列是金额
                        description = row[9].value if len(row) > 9 and row[9].value else ''
                        amount = safe_float(row[10].value) if len(row) > 10 else 0
                        category = description  # 使用摘要作为分类
                        
                        if amount > 0:
                            transaction = Transaction(
                                date=date,
                                type='支出',
                                category=str(category),
                                amount=amount,
                                description=str(description),
                                month=date.month,
                                year=date.year
                            )
                            transactions.append(transaction)
        
        elif sheet_name in ['一月', '二月', '三月', '四月', '五月', '六月', 
                           '七月', '八月', '九月', '十月', '十一月', '十二月',
                           '26一月', '26二月', '26三月']:
            for row in ws.iter_rows(min_row=4):
                if len(row) > 0 and row[0].value:
                    date = parse_date(row[0].value)
                    if date:
                        # 修正：第2列是摘要，第3列是金额
                        description = row[1].value if len(row) > 1 and row[1].value else ''
                        amount = safe_float(row[2].value) if len(row) > 2 else 0
                        category = description  # 使用摘要作为分类
                        
                        if amount > 0:
                            transaction = Transaction(
                                date=date,
                                type='支出',
                                category=str(category),
                                amount=amount,
                                description=str(description),
                                month=date.month,
                                year=date.year
                            )
                            transactions.append(transaction)
                
                if len(row) > 4 and row[4].value:
                    date = parse_date(row[4].value)
                    if date:
                        # 修正：第6列是摘要，第7列是金额
                        description = row[5].value if len(row) > 5 and row[5].value else ''
                        amount = safe_float(row[6].value) if len(row) > 6 else 0
                        category = description  # 使用摘要作为分类
                        
                        if amount > 0:
                            transaction = Transaction(
                                date=date,
                                type='入账',
                                category=str(category),
                                amount=amount,
                                description=str(description),
                                month=date.month,
                                year=date.year
                            )
                            transactions.append(transaction)
        
        elif sheet_name == '借款':
            pass
    
    wb.close()
    
    return {
        'transactions': transactions,
        'loans': [],
        'distributions': []
    }
