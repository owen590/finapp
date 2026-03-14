import openpyxl
import json

wb = openpyxl.load_workbook('e:\\账务系统\\恒驰泵业.xlsx')

print("=== Excel文件结构分析 ===\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"\n{'='*50}")
    print(f"Sheet: {sheet_name}")
    print(f"{'='*50}")
    
    # 获取有数据的行数和列数
    max_row = ws.max_row
    max_col = ws.max_column
    
    print(f"数据范围: {max_row} 行 x {max_col} 列")
    
    # 读取前20行数据
    print("\n前20行数据:")
    for row_idx, row in enumerate(ws.iter_rows(max_row=20), 1):
        row_data = [cell.value for cell in row]
        # 过滤掉None值
        non_null_data = [(i+1, val) for i, val in enumerate(row_data) if val is not None]
        if non_null_data:
            print(f"Row {row_idx}: {non_null_data}")
    
    print(f"\n总共有 {max_row} 行数据")

wb.close()
